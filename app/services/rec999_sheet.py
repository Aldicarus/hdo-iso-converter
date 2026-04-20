"""
rec999_sheet.py — Lectura live del Google Sheet de R3S3T_9999 con
información de qué películas son convertibles a CMv4.0 y cuáles no.

El sheet tiene tres secciones por fila:
  - Izquierda (col 0-4): películas NO factibles (BD-FEL exclusivas o con
    problemas detallados en la columna Notes)
  - Derecha  (col 6-11): películas convertibles via workflow 2-3
    (transferencia del bloque CMv4.0 desde un WEB-DL al RPU P7 del BD)
  - Muy a la derecha: encoder groups a evitar (ignorado aquí)

La lectura se hace live sin API key usando el endpoint público
`/export?format=csv&gid=...` de Google Sheets (el sheet tiene permisos
"cualquiera con el enlace puede ver"). Caché en memoria + disco como
fallback.
"""
from __future__ import annotations

import csv
import io
import logging
import os
import re
import time
from pathlib import Path

import httpx
from pydantic import BaseModel

from services.settings_store import (
    get_google_api_key,
    get_cmv40_sheet_id_gid,
)

_logger = logging.getLogger(__name__)

# El ID y GID del sheet se leen en runtime desde settings (con default público
# al sheet DoviTools oficial). El usuario puede cambiarlo en la pantalla de
# Configuración y se aplica sin reiniciar.


def _sheet_ids() -> tuple[str, str]:
    """(sheet_id, gid) actual. Consultar en runtime."""
    return get_cmv40_sheet_id_gid()


def _sheet_csv_url() -> str:
    sid, gid = _sheet_ids()
    return (f"https://docs.google.com/spreadsheets/d/{sid}"
            f"/export?format=csv&gid={gid}")


def _sheet_html_url() -> str:
    sid, gid = _sheet_ids()
    return (f"https://docs.google.com/spreadsheets/d/{sid}"
            f"/gviz/tq?tqx=out:html&gid={gid}")


def _sheet_xlsx_url() -> str:
    sid, gid = _sheet_ids()
    return (f"https://docs.google.com/spreadsheets/d/{sid}"
            f"/export?format=xlsx&gid={gid}")

CONFIG_DIR = Path(os.environ.get("CONFIG_DIR", "/config"))
CACHE_PATH = CONFIG_DIR / "rec999_sheet_cache.csv"
CACHE_TTL_SECONDS = 3600  # 1h

# Año entre 1950 y 2099 — elimina falsos positivos con números random
_YEAR_RE = re.compile(r"\b(19[5-9]\d|20\d{2})\b")

# Frases que identifican filas de cabecera o instrucciones (no películas)
_NON_MOVIE_MARKERS = (
    "incorrect p8", "when fel is not", "generate dolby", "restore cmv",
    "how to make proper", "dolby vision test", "hdr10", "fel vs",
    "dv metadata", "fel bluray", "best player", "questions",
    "aspect ratio", "encoder groups",
)


class RecommendationRow(BaseModel):
    """Una fila parseada del sheet."""
    feasible: bool
    title_raw: str
    title_normalized: str
    year: int | None = None
    dv_source: str = ""           # 'BD FEL', 'iTunes', 'DSNP', 'MA'…
    sync_offset: str = ""         # '(+24)', '0', '(-8 T280/B280)'…
    sync_offset_frames: int | None = None
    comparisons: str = ""         # primera sub-columna de Comparisons (HDR COMP, nits…)
    comparisons_2: str = ""       # segunda sub-columna de Comparisons (L1, plot…)
    notes: str = ""               # razón detallada / workflow
    # Hyperlinks por columna (solo disponibles si la hoja se leyó via
    # Sheets API v4 o XLSX+openpyxl; vacío si se leyó via CSV export)
    title_link: str = ""
    sync_link: str = ""
    dv_source_link: str = ""
    comparisons_link: str = ""
    comparisons_2_link: str = ""
    notes_link: str = ""


# Caché en memoria
_cache_rows: list[RecommendationRow] | None = None
_cache_fetched_at: float = 0.0
_cache_source: str = "none"          # 'api' | 'csv' | 'disk' | 'none'
_sheets_api_error: str = ""          # último error de Sheets API (para UI)


def get_cache_status() -> dict:
    """Estado del último fetch del sheet (para exponer al frontend)."""
    return {
        "source": _cache_source,
        "fetched_at": _cache_fetched_at,
        "rows": len(_cache_rows) if _cache_rows is not None else 0,
        "sheets_api_error": _sheets_api_error,
    }


def invalidate_cache() -> None:
    """Fuerza re-fetch en la próxima llamada (útil al cambiar Google key)."""
    global _cache_rows, _cache_fetched_at, _cache_source
    _cache_rows = None
    _cache_fetched_at = 0.0
    _cache_source = "none"


def _normalize_title(raw: str) -> str:
    """Slug: minúsculas, sin puntuación, sin año, espacios colapsados."""
    s = raw.lower()
    s = re.sub(r"\[[^\]]+\]", " ", s)
    s = re.sub(r"[\._\-/:]+", " ", s)
    s = _YEAR_RE.sub("", s)
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _extract_year(raw: str) -> int | None:
    """Devuelve el año de estreno. Cuando el título contiene varios años
    (p.ej. 'Blade.Runner.2049.2017', '2001.A.Space.Odyssey.1968'), el año
    real es el ÚLTIMO — los anteriores forman parte del título."""
    matches = _YEAR_RE.findall(raw)
    return int(matches[-1]) if matches else None


def _parse_offset_frames(s: str) -> int | None:
    """De '(+24)', '(-8 T280/B280)', '0' saca el primer entero con signo."""
    if not s:
        return None
    m = re.search(r"([+-]?\d+)", s)
    return int(m.group(1)) if m else None


def _is_instructional(title: str) -> bool:
    low = title.lower()
    return any(marker in low for marker in _NON_MOVIE_MARKERS)


# Regex para extraer URLs que aparecen como texto plano en una celda
# (no como <a href>). Muy común en los Notes de REC_9999, tipo
# "L5 edits https://justpaste.it/bhv73".
_URL_RE = re.compile(r"https?://[^\s<>\"']+")


def _extract_first_url(text: str) -> str:
    m = _URL_RE.search(text or "")
    return m.group(0) if m else ""


def _enrich_cell_link(cell: dict) -> dict:
    """Si la celda no tiene link explícito (ej. por `<a href>`), busca
    una URL plana en su texto. Muchas celdas del sheet tienen URLs como
    texto, no como hyperlinks formales."""
    if not cell.get("link"):
        url = _extract_first_url(cell.get("text", ""))
        if url:
            return {"text": cell.get("text", ""), "link": url}
    return cell


def _parse_row(row: list[dict]) -> list[RecommendationRow]:
    """Devuelve 0, 1 o 2 RecommendationRow por fila. Cada celda es
    `{"text": str, "link": str}` (link puede ser "" si no hay o si se
    leyó vía CSV)."""
    out: list[RecommendationRow] = []
    if len(row) < 5:
        return out

    def col(i: int) -> dict:
        raw = (row[i] if i < len(row) else {"text": "", "link": ""})
        return _enrich_cell_link(raw)

    # Izquierda: NO factible
    # cols: 0=title, 1=dv_source, 2+3=comparisons (2 sub-cols), 4=notes
    left = col(0)
    left_title = left["text"].strip()
    if left_title and _extract_year(left_title) and not _is_instructional(left_title):
        out.append(RecommendationRow(
            feasible=False,
            title_raw=left_title,
            title_link=left["link"],
            title_normalized=_normalize_title(left_title),
            year=_extract_year(left_title),
            dv_source=col(1)["text"].strip(),
            dv_source_link=col(1)["link"],
            comparisons=col(2)["text"].strip(),
            comparisons_link=col(2)["link"],
            comparisons_2=col(3)["text"].strip(),
            comparisons_2_link=col(3)["link"],
            notes=col(4)["text"].strip(),
            notes_link=col(4)["link"],
        ))

    # Derecha: factible
    # cols: 6=title, 7=sync, 8=dv_source, 9+10=comparisons (2 sub-cols), 11=notes
    right = col(6)
    right_title = right["text"].strip()
    if right_title and _extract_year(right_title) and right_title.upper() != "OLD":
        sync_text = col(7)["text"]
        out.append(RecommendationRow(
            feasible=True,
            title_raw=right_title,
            title_link=right["link"],
            title_normalized=_normalize_title(right_title),
            year=_extract_year(right_title),
            sync_offset=sync_text.strip(),
            sync_link=col(7)["link"],
            sync_offset_frames=_parse_offset_frames(sync_text),
            dv_source=col(8)["text"].strip(),
            dv_source_link=col(8)["link"],
            comparisons=col(9)["text"].strip(),
            comparisons_link=col(9)["link"],
            comparisons_2=col(10)["text"].strip(),
            comparisons_2_link=col(10)["link"],
            notes=col(11)["text"].strip(),
            notes_link=col(11)["link"],
        ))

    # Extra derecha: sección "Not Sure! / probably ok" (cols 13-18).
    # Mismo layout que la factible pero son casos con incertidumbre (MDL
    # mismatch, L5 edits, etc.). Los tratamos como feasible=True porque
    # REC_9999 dice "probably ok" → son convertibles con caveats.
    # El header de col 13 es "BHDstudio" (encoder groups a evitar) — se
    # filtra solo porque no tiene año.
    extra = col(13)
    extra_title = extra["text"].strip()
    if (extra_title and _extract_year(extra_title)
            and not _is_instructional(extra_title)
            and extra_title.upper() != "OLD"):
        sync_text = col(14)["text"]
        # cols: 13=title, 14=sync, 15=dv_source, 16+17=comparisons (2 sub-cols), 18=notes
        out.append(RecommendationRow(
            feasible=True,
            title_raw=extra_title,
            title_link=extra["link"],
            title_normalized=_normalize_title(extra_title),
            year=_extract_year(extra_title),
            sync_offset=sync_text.strip(),
            sync_link=col(14)["link"],
            sync_offset_frames=_parse_offset_frames(sync_text),
            dv_source=col(15)["text"].strip(),
            dv_source_link=col(15)["link"],
            comparisons=col(16)["text"].strip(),
            comparisons_link=col(16)["link"],
            comparisons_2=col(17)["text"].strip(),
            comparisons_2_link=col(17)["link"],
            notes=col(18)["text"].strip(),
            notes_link=col(18)["link"],
        ))

    return out


def parse_csv_text(csv_text: str) -> list[RecommendationRow]:
    """Parsea el CSV completo (sin links) y devuelve todas las filas."""
    rows: list[RecommendationRow] = []
    reader = csv.reader(io.StringIO(csv_text))
    for i, row in enumerate(reader):
        if i == 0:  # header
            continue
        # Convierte strings planos a dicts {text, link} para uniformar
        cells = [{"text": c, "link": ""} for c in row]
        rows.extend(_parse_row(cells))
    return rows


def parse_api_rows(api_rows: list[list[dict]]) -> list[RecommendationRow]:
    """Parsea la estructura que devuelve Sheets API v4 (con hyperlinks)."""
    result: list[RecommendationRow] = []
    for i, row in enumerate(api_rows):
        if i == 0:  # header
            continue
        result.extend(_parse_row(row))
    return result


async def _fetch_csv() -> str:
    async with httpx.AsyncClient(timeout=20.0, follow_redirects=True) as client:
        resp = await client.get(_sheet_csv_url())
        resp.raise_for_status()
        return resp.text


# HTML export vía gviz — funciona con sheets XLSX importados (donde Sheets
# API v4 falla con "This operation is not supported for this document").
# URL se construye en runtime con _sheet_html_url() para respetar cambios
# en la configuración del sheet desde la UI.


def _parse_gviz_html(html: str) -> list[list[dict]]:
    """Parsea la tabla HTML que devuelve `/gviz/tq?tqx=out:html`.
    Cada celda se devuelve como `{"text": str, "link": str}`. Si la celda
    tiene un `<a>`, se extrae su href (desempaquetando el wrapper
    `https://www.google.com/url?q=...` que pone a veces Google)."""
    from html.parser import HTMLParser
    import html as _htmllib
    import urllib.parse

    def _unwrap(href: str) -> str:
        if href.startswith("https://www.google.com/url?q="):
            try:
                qs = urllib.parse.urlparse(href).query
                params = urllib.parse.parse_qs(qs)
                return params.get("q", [href])[0]
            except Exception:
                return href
        return href

    class _P(HTMLParser):
        def __init__(self) -> None:
            super().__init__(convert_charrefs=True)
            self.rows: list[list[dict]] = []
            self._row: list[dict] | None = None
            self._in_cell = False
            self._text: list[str] = []
            self._href = ""

        def handle_starttag(self, tag, attrs):
            if tag == "tr":
                self._row = []
            elif tag in ("td", "th"):
                self._in_cell = True
                self._text = []
                self._href = ""
            elif tag == "a" and self._in_cell and not self._href:
                for k, v in attrs:
                    if k == "href" and v:
                        self._href = _unwrap(v)
                        break
            elif tag == "br" and self._in_cell:
                self._text.append("\n")

        def handle_endtag(self, tag):
            if tag in ("td", "th"):
                self._in_cell = False
                text = "".join(self._text).strip()
                if self._row is not None:
                    self._row.append({"text": text, "link": self._href})
                self._text = []
                self._href = ""
            elif tag == "tr":
                if self._row is not None:
                    self.rows.append(self._row)
                    self._row = None

        def handle_data(self, data):
            if self._in_cell:
                self._text.append(data)

    p = _P()
    p.feed(html)
    return p.rows


# XLSX export — única vía que preserva rich-text hyperlinks (los que el
# usuario crea seleccionando texto y pulsando "insertar enlace", sin usar
# la fórmula =HYPERLINK). gviz HTML y CSV los pierden. URL se construye
# en runtime con _sheet_xlsx_url().


async def _fetch_sheet_xlsx() -> list[list[dict]] | None:
    """Descarga el sheet como XLSX y lo parsea con openpyxl, preservando
    los hyperlinks rich-text de cada celda."""
    try:
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            resp = await client.get(_sheet_xlsx_url())
            if resp.status_code != 200:
                _logger.info("XLSX export no accesible (%d)", resp.status_code)
                return None
            xlsx_bytes = resp.content
    except Exception as e:
        _logger.warning("XLSX fetch falló: %s", e)
        return None

    try:
        import io as _io
        from openpyxl import load_workbook
    except ImportError:
        _logger.warning("openpyxl no instalado — no puedo extraer rich-text links")
        return None

    try:
        wb = load_workbook(_io.BytesIO(xlsx_bytes), read_only=False, data_only=True)
    except Exception as e:
        _logger.warning("openpyxl load_workbook falló: %s", e)
        return None

    # openpyxl no expone el `sheetId` (gid) de Google; busco la pestaña
    # cuyo header encaje con "Title,DV Source,Comparisons…". Fallback a
    # la primera pestaña si no hay match claro.
    target = None
    for name in wb.sheetnames:
        ws = wb[name]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not row1:
            continue
        cells = [str(c).strip().lower() if c is not None else "" for c in row1]
        if cells and cells[0] == "title" and any("dv source" in c for c in cells):
            target = ws
            break
    if target is None and wb.sheetnames:
        target = wb[wb.sheetnames[0]]
    if target is None:
        return None

    rows: list[list[dict]] = []
    for row in target.iter_rows():
        cells: list[dict] = []
        for cell in row:
            text_val = cell.value
            if text_val is None:
                text = ""
            elif isinstance(text_val, str):
                text = text_val
            else:
                text = str(text_val)
            link = ""
            try:
                if cell.hyperlink:
                    # cell.hyperlink.target es el URL en rich-text e =HYPERLINK
                    link = cell.hyperlink.target or cell.hyperlink.location or ""
            except Exception:
                pass
            cells.append({"text": text, "link": link})
        rows.append(cells)
    return rows


async def _fetch_sheet_html() -> list[list[dict]] | None:
    """Export HTML público (gviz) con hyperlinks preservados. Funciona en
    sheets XLSX importados donde Sheets API v4 no aplica."""
    try:
        async with httpx.AsyncClient(timeout=25.0, follow_redirects=True) as client:
            resp = await client.get(_sheet_html_url())
            if resp.status_code != 200:
                _logger.info("gviz HTML no accesible (%d)", resp.status_code)
                return None
            rows = _parse_gviz_html(resp.text)
            return rows if rows else None
    except Exception as e:
        _logger.warning("gviz HTML fetch falló: %s", e)
        return None


def _parse_sheets_api_error(resp) -> str:
    """Extrae un mensaje human-readable del error 4xx de Sheets API."""
    try:
        err = (resp.json().get("error") or {})
        msg = err.get("message", "")
        status = err.get("status", "")
        # Mensaje típico cuando Sheets API no está habilitada:
        # "Google Sheets API has not been used in project ... before or it is
        # disabled. Enable it by visiting https://console.developers.google.com/apis/api/sheets.googleapis.com/overview?project=..."
        if "has not been used" in msg or "disabled" in msg:
            return ("Google Sheets API no está habilitada en tu proyecto de "
                    "Google Cloud. Actívala en "
                    "https://console.cloud.google.com/apis/library/sheets.googleapis.com")
        if status == "PERMISSION_DENIED":
            return f"Acceso denegado por Google: {msg}"
        return msg or f"HTTP {resp.status_code}"
    except Exception:
        return f"HTTP {resp.status_code}"


async def _fetch_sheet_api(api_key: str) -> list[list[dict]] | None:
    """Sheets API v4 con hyperlinks por celda. Devuelve filas estructuradas
    `[[{text,link}, ...], ...]` o None si falla (sheet privado, API no
    habilitada, etc.). Actualiza `_sheets_api_error` con el motivo."""
    global _sheets_api_error
    _sheets_api_error = ""
    sheet_id, sheet_gid = _sheet_ids()
    try:
        async with httpx.AsyncClient(timeout=25.0, follow_redirects=True) as client:
            # 1. Resolver nombre de pestaña desde el GID
            meta_resp = await client.get(
                f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}",
                params={"key": api_key,
                        "fields": "sheets(properties(title,sheetId))"},
            )
            if meta_resp.status_code != 200:
                _sheets_api_error = _parse_sheets_api_error(meta_resp)
                _logger.info("Sheets API v4 metadatos no accesibles (%d): %s",
                             meta_resp.status_code, _sheets_api_error)
                return None
            sheets_meta = meta_resp.json().get("sheets", [])
            sheet_name: str | None = None
            for s in sheets_meta:
                if str(s.get("properties", {}).get("sheetId")) == str(sheet_gid):
                    sheet_name = s["properties"].get("title")
                    break
            if not sheet_name:
                _sheets_api_error = f"Pestaña con gid={sheet_gid} no encontrada"
                _logger.warning(_sheets_api_error)
                return None

            # 2. Grid data con formattedValue + hyperlinks + rich-text links
            fields = (
                "sheets(data(rowData(values("
                "formattedValue,hyperlink,"
                "textFormatRuns(format(link(uri)),startIndex)"
                "))))"
            )
            grid_resp = await client.get(
                f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}",
                params={
                    "key": api_key,
                    "ranges": sheet_name,
                    "includeGridData": "true",
                    "fields": fields,
                },
            )
            if grid_resp.status_code != 200:
                _sheets_api_error = _parse_sheets_api_error(grid_resp)
                _logger.info("Sheets API v4 grid-data no accesible (%d): %s",
                             grid_resp.status_code, _sheets_api_error)
                return None
            data = grid_resp.json()
    except Exception as e:
        _sheets_api_error = f"Error de red: {e}"
        _logger.warning("Sheets API v4 falló: %s", e)
        return None

    rows: list[list[dict]] = []
    for sheet in data.get("sheets", []):
        for grid in sheet.get("data", []):
            for row in grid.get("rowData", []):
                cells: list[dict] = []
                for cell in row.get("values", []):
                    text = cell.get("formattedValue", "") or ""
                    link = cell.get("hyperlink", "") or ""
                    if not link:
                        # Rich-text runs (link manual a parte del texto)
                        for run in (cell.get("textFormatRuns") or []):
                            uri = (run.get("format", {})
                                        .get("link", {})
                                        .get("uri"))
                            if uri:
                                link = uri
                                break
                    cells.append({"text": text, "link": link})
                rows.append(cells)
    return rows


async def get_recommendations(force_refresh: bool = False) -> list[RecommendationRow]:
    """Devuelve las filas parseadas.

    Prioridad de fetch:
      1. Sheets API v4 (si hay Google API key) — preserva hyperlinks
      2. CSV export (público, sin auth) — solo texto
      3. Caché en disco (última copia buena)
    """
    global _cache_rows, _cache_fetched_at, _cache_source

    now = time.time()
    if (not force_refresh and _cache_rows is not None
            and now - _cache_fetched_at < CACHE_TTL_SECONDS):
        return _cache_rows

    # Intento 1: Sheets API v4 con hyperlinks (solo funciona en sheets
    # nativos, no en XLSX importados; requiere Google key)
    google_key = get_google_api_key()
    if google_key:
        api_rows = await _fetch_sheet_api(google_key)
        if api_rows:
            parsed = parse_api_rows(api_rows)
            if parsed:
                _cache_rows = parsed
                _cache_fetched_at = now
                _cache_source = "api"
                _logger.info("REC_9999 sheet: %d filas via Sheets API v4 "
                             "(con hyperlinks)", len(parsed))
                return parsed

    # Intento 2: export XLSX + openpyxl — la única vía que preserva
    # hyperlinks rich-text (los que REC_9999 crea seleccionando texto +
    # insertar enlace, muy habituales en su sheet). Sheets API v4, gviz
    # HTML y CSV los PIERDEN. No requiere auth.
    xlsx_rows = await _fetch_sheet_xlsx()
    if xlsx_rows:
        parsed = parse_api_rows(xlsx_rows)
        if parsed:
            _cache_rows = parsed
            _cache_fetched_at = now
            _cache_source = "xlsx"
            _logger.info("REC_9999 sheet: %d filas via XLSX export "
                         "(con hyperlinks rich-text)", len(parsed))
            return parsed

    # Intento 3: export HTML (gviz) — preserva =HYPERLINK() pero no
    # rich-text links. Útil si openpyxl no está instalado.
    html_rows = await _fetch_sheet_html()
    if html_rows:
        parsed = parse_api_rows(html_rows)
        if parsed:
            _cache_rows = parsed
            _cache_fetched_at = now
            _cache_source = "html"
            _logger.info("REC_9999 sheet: %d filas via HTML export "
                         "(solo hyperlinks formales)", len(parsed))
            return parsed

    # Intento 4: CSV export público
    csv_text: str | None = None
    live_csv_ok = False
    try:
        csv_text = await _fetch_csv()
        live_csv_ok = True
        try:
            CONFIG_DIR.mkdir(parents=True, exist_ok=True)
            CACHE_PATH.write_text(csv_text, encoding="utf-8")
        except OSError as e:
            _logger.warning("No pude persistir cache del sheet: %s", e)
    except Exception as e:
        _logger.warning("Fetch CSV del sheet REC_9999 falló: %s", e)

    # Intento 3: último CSV bueno en disco
    source = "csv" if live_csv_ok else "disk"
    if csv_text is None and CACHE_PATH.exists():
        try:
            csv_text = CACHE_PATH.read_text(encoding="utf-8")
            source = "disk"
        except OSError:
            csv_text = None

    if not csv_text:
        _logger.error("Sin datos del sheet REC_9999 (ni API, ni CSV, ni cache)")
        _cache_rows = []
        _cache_fetched_at = now
        _cache_source = "none"
        return []

    _cache_rows = parse_csv_text(csv_text)
    _cache_fetched_at = now
    _cache_source = source
    _logger.info("REC_9999 sheet: %d filas via %s (sin hyperlinks)",
                 len(_cache_rows), source.upper())
    return _cache_rows
